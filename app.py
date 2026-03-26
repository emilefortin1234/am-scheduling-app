
import io
import random
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook

st.set_page_config(page_title="Scheduling en fabrication additive", layout="wide")

DEFAULT_MATERIAL_NAMES = ["PLA", "ABS", "RESIN", "NYLON", "PETG", "TPU", "PEEK", "ALU_FILL"]


def init_state():
    if "materials_df" not in st.session_state:
        st.session_state.materials_df = pd.DataFrame(
            {"material_id": [1, 2, 3], "material_name": ["PLA", "ABS", "RESIN"]}
        )
    if "machines_df" not in st.session_state:
        st.session_state.machines_df = pd.DataFrame(
            {
                "machine_id": [1, 2],
                "VT": [0.030864, 0.030864],
                "HT": [0.7, 0.7],
                "SET": [1.0, 1.2],
                "MA": [800.0, 1200.0],
                "MH": [32.0, 40.0],
            }
        )
    if "parts_df" not in st.session_state:
        st.session_state.parts_df = pd.DataFrame(
            {
                "part_id": list(range(1, 11)),
                "h": [6.9, 26.04, 15.97, 17.04, 27.94, 12.0, 8.0, 11.5, 9.8, 14.2],
                "a": [209.06, 550.11, 23.63, 99.53, 56.85, 120.0, 210.0, 160.0, 95.0, 180.0],
                "v": [826.08, 952.60, 71.91, 703.08, 272.92, 300.0, 1100.0, 760.0, 430.0, 980.0],
                "material_name": ["PLA", "ABS", "PLA", "RESIN", "ABS", "PLA", "RESIN", "PLA", "ABS", "PLA"],
            }
        )
    if "compat_df" not in st.session_state:
        st.session_state.compat_df = pd.DataFrame(
            [[1, 1, 0], [1, 1, 1]],
            index=["M1", "M2"],
            columns=["PLA", "ABS", "RESIN"],
        )
    if "jn" not in st.session_state:
        st.session_state.jn = 4


def sync_all():
    materials_df = st.session_state.materials_df.copy()
    machines_df = st.session_state.machines_df.copy()
    parts_df = st.session_state.parts_df.copy()
    compat_df = st.session_state.compat_df.copy()

    if materials_df.empty:
        materials_df = pd.DataFrame({"material_id": [1], "material_name": ["PLA"]})
    materials_df = materials_df.dropna(subset=["material_name"]).copy()
    materials_df["material_name"] = materials_df["material_name"].astype(str).str.strip()
    materials_df = materials_df[materials_df["material_name"] != ""].drop_duplicates(subset=["material_name"]).reset_index(drop=True)
    if materials_df.empty:
        materials_df = pd.DataFrame({"material_id": [1], "material_name": ["PLA"]})
    materials_df["material_id"] = range(1, len(materials_df) + 1)
    valid_materials = materials_df["material_name"].tolist()

    if machines_df.empty:
        machines_df = pd.DataFrame(columns=["machine_id", "VT", "HT", "SET", "MA", "MH"])
    else:
        machines_df = machines_df.dropna(how="all").copy()
        machines_df["machine_id"] = range(1, len(machines_df) + 1)

    if parts_df.empty:
        parts_df = pd.DataFrame(columns=["part_id", "h", "a", "v", "material_name"])
    else:
        parts_df = parts_df.dropna(how="all").copy()
        parts_df["part_id"] = range(1, len(parts_df) + 1)
        parts_df["material_name"] = parts_df["material_name"].astype(str).str.strip()
        default_mat = valid_materials[0]
        parts_df.loc[~parts_df["material_name"].isin(valid_materials), "material_name"] = default_mat

    machine_labels = [f"M{mid}" for mid in machines_df["machine_id"].tolist()]
    if compat_df.empty:
        compat_df = pd.DataFrame(0, index=machine_labels, columns=valid_materials)
    compat_df = compat_df.reindex(index=machine_labels, columns=valid_materials, fill_value=0)
    compat_df = compat_df.fillna(0).astype(int)

    st.session_state.materials_df = materials_df.reset_index(drop=True)
    st.session_state.machines_df = machines_df.reset_index(drop=True)
    st.session_state.parts_df = parts_df.reset_index(drop=True)
    st.session_state.compat_df = compat_df


def generate_materials(n_materials: int):
    names = DEFAULT_MATERIAL_NAMES[:n_materials]
    if len(names) < n_materials:
        names += [f"MAT_{i}" for i in range(len(names) + 1, n_materials + 1)]
    st.session_state.materials_df = pd.DataFrame({"material_id": list(range(1, n_materials + 1)), "material_name": names})
    sync_all()


def generate_machines(n_machines: int, ma_min: float, ma_max: float, mh_min: float, mh_max: float):
    rows = []
    for m in range(1, n_machines + 1):
        rows.append(
            {
                "machine_id": m,
                "VT": round(random.uniform(0.02, 0.05), 6),
                "HT": round(random.uniform(0.5, 1.0), 3),
                "SET": round(random.uniform(0.8, 1.5), 3),
                "MA": round(random.uniform(ma_min, ma_max), 2),
                "MH": round(random.uniform(mh_min, mh_max), 2),
            }
        )
    st.session_state.machines_df = pd.DataFrame(rows)
    sync_all()


def generate_parts(n_parts: int, h_min: float, h_max: float, a_min: float, a_max: float, v_min: float, v_max: float):
    mats = st.session_state.materials_df["material_name"].tolist() or ["PLA"]
    rows = []
    for i in range(1, n_parts + 1):
        rows.append(
            {
                "part_id": i,
                "h": round(random.uniform(h_min, h_max), 2),
                "a": round(random.uniform(a_min, a_max), 2),
                "v": round(random.uniform(v_min, v_max), 2),
                "material_name": random.choice(mats),
            }
        )
    st.session_state.parts_df = pd.DataFrame(rows)
    sync_all()


def generate_compatibility():
    machines_df = st.session_state.machines_df
    materials_df = st.session_state.materials_df
    parts_df = st.session_state.parts_df

    machine_labels = [f"M{mid}" for mid in machines_df["machine_id"].tolist()]
    material_names = materials_df["material_name"].tolist()
    compat = pd.DataFrame(0, index=machine_labels, columns=material_names)

    if not machine_labels or not material_names:
        st.session_state.compat_df = compat
        return

    for m in machine_labels:
        k = random.randint(1, max(1, len(material_names)))
        chosen = random.sample(material_names, k=k)
        for mat in chosen:
            compat.loc[m, mat] = 1

    used_mats = sorted(set(parts_df["material_name"].tolist())) if not parts_df.empty else material_names
    for mat in used_mats:
        compat.loc[random.choice(machine_labels), mat] = 1

    for m in machine_labels:
        if compat.loc[m].sum() == 0:
            compat.loc[m, random.choice(material_names)] = 1

    st.session_state.compat_df = compat.astype(int)
    sync_all()


def validate_dataframes(parts_df, machines_df, compat_df, jn):
    errors, warnings = [], []

    if machines_df.empty:
        errors.append("Aucune machine n'est définie.")
    if parts_df.empty:
        errors.append("Aucune pièce n'est définie.")
    if compat_df.empty:
        errors.append("La matrice de compatibilité est vide.")
    if errors:
        return errors, warnings

    used_materials = sorted(set(parts_df["material_name"].tolist()))
    max_batches = len(machines_df) * int(jn)
    if max_batches < len(used_materials):
        warnings.append(f"Le nombre maximal de batches ({max_batches}) est inférieur au nombre de matériaux utilisés ({len(used_materials)}).")

    for mat in used_materials:
        if mat not in compat_df.columns or compat_df[mat].sum() == 0:
            errors.append(f"Le matériau '{mat}' est utilisé, mais aucune machine n'est compatible.")

    for _, p in parts_df.iterrows():
        admissible = []
        for _, m in machines_df.iterrows():
            machine_label = f"M{int(m['machine_id'])}"
            mat = p["material_name"]
            is_compat = machine_label in compat_df.index and mat in compat_df.columns and int(compat_df.loc[machine_label, mat]) == 1
            if is_compat and float(m["MH"]) >= float(p["h"]) and float(m["MA"]) >= float(p["a"]):
                admissible.append(machine_label)
        if not admissible:
            errors.append(f"Pièce P{int(p['part_id'])}: aucune machine admissible (matériau={p['material_name']}, h={p['h']}, a={p['a']}).")
        elif len(admissible) == 1:
            warnings.append(f"Pièce P{int(p['part_id'])}: une seule machine possible ({admissible[0]}).")

    return errors, warnings


def schedule_heuristic(parts_df, machines_df, compat_df, jn):
    batches_by_machine = {int(mid): [] for mid in machines_df["machine_id"].tolist()}
    unassigned = []
    machines_map = {int(row["machine_id"]): row.to_dict() for _, row in machines_df.iterrows()}

    def admissible_machine_ids(part_row):
        out = []
        for _, m in machines_df.iterrows():
            mid = int(m["machine_id"])
            mlabel = f"M{mid}"
            mat = part_row["material_name"]
            is_compat = mlabel in compat_df.index and mat in compat_df.columns and int(compat_df.loc[mlabel, mat]) == 1
            if is_compat and float(m["MH"]) >= float(part_row["h"]) and float(m["MA"]) >= float(part_row["a"]):
                out.append(mid)
        return out

    parts_sorted = parts_df.copy()
    parts_sorted["n_admissible"] = parts_sorted.apply(lambda row: len(admissible_machine_ids(row)), axis=1)
    parts_sorted = parts_sorted.sort_values(["n_admissible", "a", "h"], ascending=[True, False, False])

    for _, p in parts_sorted.iterrows():
        feasible_machines = admissible_machine_ids(p)
        if not feasible_machines:
            unassigned.append(int(p["part_id"]))
            continue

        placed = False
        candidates = []
        for mid in feasible_machines:
            machine = machines_map[mid]
            for batch in batches_by_machine[mid]:
                if batch["material_name"] != p["material_name"]:
                    continue
                new_area = batch["area_total"] + float(p["a"])
                new_height = max(batch["height_max"], float(p["h"]))
                if new_area <= float(machine["MA"]) and new_height <= float(machine["MH"]):
                    projected_pt = float(machine["SET"]) + float(machine["VT"]) * (batch["volume_total"] + float(p["v"])) + float(machine["HT"]) * new_height
                    candidates.append((projected_pt, len(batch["parts"]), mid, batch))
        if candidates:
            _, _, _, chosen = min(candidates, key=lambda x: (x[0], x[1]))
            chosen["parts"].append(int(p["part_id"]))
            chosen["area_total"] += float(p["a"])
            chosen["volume_total"] += float(p["v"])
            chosen["height_max"] = max(chosen["height_max"], float(p["h"]))
            chosen["part_rows"].append(p.to_dict())
            placed = True

        if not placed:
            candidates = []
            for mid in feasible_machines:
                if len(batches_by_machine[mid]) >= int(jn):
                    continue
                machine = machines_map[mid]
                pt = float(machine["SET"]) + float(machine["VT"]) * float(p["v"]) + float(machine["HT"]) * float(p["h"])
                candidates.append((len(batches_by_machine[mid]), pt, mid))
            if candidates:
                _, _, mid = min(candidates, key=lambda x: (x[0], x[1]))
                batches_by_machine[mid].append(
                    {
                        "machine_id": mid,
                        "material_name": p["material_name"],
                        "parts": [int(p["part_id"])],
                        "part_rows": [p.to_dict()],
                        "area_total": float(p["a"]),
                        "volume_total": float(p["v"]),
                        "height_max": float(p["h"]),
                    }
                )
                placed = True

        if not placed:
            unassigned.append(int(p["part_id"]))

    batch_rows, gantt_rows, comp_rows, machine_summary = [], [], [], []
    cmax = 0.0

    for _, m in machines_df.iterrows():
        mid = int(m["machine_id"])
        time_cursor = 0.0
        batches = batches_by_machine[mid]
        for j, batch in enumerate(batches, start=1):
            pt = float(m["SET"]) + float(m["VT"]) * batch["volume_total"] + float(m["HT"]) * batch["height_max"]
            start = time_cursor
            end = start + pt
            batch_id = f"M{mid}-B{j}"
            time_cursor = end
            cmax = max(cmax, end)

            batch_rows.append(
                {
                    "batch_id": batch_id,
                    "machine": f"M{mid}",
                    "batch_index": j,
                    "material": batch["material_name"],
                    "pieces": ", ".join(f"P{pid}" for pid in batch["parts"]),
                    "nb_pieces": len(batch["parts"]),
                    "height_max": round(batch["height_max"], 2),
                    "area_total": round(batch["area_total"], 2),
                    "volume_total": round(batch["volume_total"], 2),
                    "PT": round(pt, 3),
                    "start": round(start, 3),
                    "end": round(end, 3),
                }
            )

            gantt_rows.append({"Machine": f"M{mid}", "Start": start, "Duration": pt, "Batch": batch_id, "Material": batch["material_name"]})

            seg = 0.0
            for pr in batch["part_rows"]:
                area = float(pr["a"])
                comp_rows.append({"Batch": batch_id, "Start": seg, "Duration": area, "Part": f"P{int(pr['part_id'])}", "Material": pr["material_name"]})
                seg += area

        machine_summary.append(
            {
                "machine": f"M{mid}",
                "nb_batches": len(batches),
                "nb_pieces": sum(len(b["parts"]) for b in batches),
                "charge_totale": round(time_cursor, 3),
            }
        )

    return pd.DataFrame(batch_rows), pd.DataFrame(gantt_rows), pd.DataFrame(comp_rows), pd.DataFrame(machine_summary), round(cmax, 3), unassigned


def build_opl_dat(parts_df, machines_df, compat_df, materials_df, jn):
    mat_to_id = {row["material_name"]: int(row["material_id"]) for _, row in materials_df.iterrows()}

    def vec(values):
        return "[\n" + ", ".join(map(str, values)) + "\n];"

    compat_rows = []
    for mid in machines_df["machine_id"].tolist():
        row = []
        mlabel = f"M{int(mid)}"
        for mat in materials_df["material_name"].tolist():
            row.append(int(compat_df.loc[mlabel, mat]) if (mlabel in compat_df.index and mat in compat_df.columns) else 0)
        compat_rows.append("[" + ", ".join(map(str, row)) + "]")

    comment_materials = "\n".join(f"// {int(row['material_id'])} = {row['material_name']}" for _, row in materials_df.iterrows())

    return (
        "n = " + str(len(parts_df)) + ";\n"
        + "mn = " + str(len(machines_df)) + ";\n"
        + "jn = " + str(jn) + ";\n"
        + "rn = " + str(len(materials_df)) + ";\n\n"
        + comment_materials + "\n\n"
        + "h = " + vec(parts_df["h"].round(2).tolist()) + "\n"
        + "a = " + vec(parts_df["a"].round(2).tolist()) + "\n"
        + "v = " + vec(parts_df["v"].round(2).tolist()) + "\n\n"
        + "mat = " + vec([mat_to_id[name] for name in parts_df["material_name"].tolist()]) + "\n\n"
        + "compat = [\n"
        + ",\n".join(compat_rows)
        + "\n];\n\n"
        + "VT = " + vec(machines_df["VT"].round(6).tolist()) + "\n"
        + "HT = " + vec(machines_df["HT"].round(3).tolist()) + "\n"
        + "SET = " + vec(machines_df["SET"].round(3).tolist()) + "\n"
        + "MA = " + vec(machines_df["MA"].round(2).tolist()) + "\n"
        + "MH = " + vec(machines_df["MH"].round(2).tolist()) + "\n"
    )


def build_excel_export(parts_df, machines_df, compat_df, materials_df, jn):
    wb = Workbook()

    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "AM Scheduling App - Export Excel"
    ws["A3"] = "Paramètres globaux"
    ws["A4"] = "Nombre de pièces"
    ws["B4"] = len(parts_df)
    ws["A5"] = "Nombre de machines"
    ws["B5"] = len(machines_df)
    ws["A6"] = "Nombre de matériaux"
    ws["B6"] = len(materials_df)
    ws["A7"] = "jn"
    ws["B7"] = int(jn)

    ws["D3"] = "Correspondance matériaux"
    ws["D4"] = "material_id"
    ws["E4"] = "material_name"
    row_start = 5
    for _, row in materials_df.iterrows():
        ws[f"D{row_start}"] = int(row["material_id"])
        ws[f"E{row_start}"] = row["material_name"]
        row_start += 1

    ws_mat = wb.create_sheet("Materials")
    ws_mat.append(["material_id", "material_name"])
    for _, row in materials_df.iterrows():
        ws_mat.append([int(row["material_id"]), row["material_name"]])

    ws_mach = wb.create_sheet("Machines")
    ws_mach.append(["machine_id", "VT", "HT", "SET", "MA", "MH"])
    for _, row in machines_df.iterrows():
        ws_mach.append([
            int(row["machine_id"]),
            float(row["VT"]),
            float(row["HT"]),
            float(row["SET"]),
            float(row["MA"]),
            float(row["MH"]),
        ])

    ws_parts = wb.create_sheet("Parts")
    ws_parts.append(["part_id", "h", "a", "v", "material_name"])
    for _, row in parts_df.iterrows():
        ws_parts.append([
            int(row["part_id"]),
            float(row["h"]),
            float(row["a"]),
            float(row["v"]),
            row["material_name"],
        ])

    ws_comp = wb.create_sheet("Compatibility")
    comp_headers = ["machine"] + list(compat_df.columns)
    ws_comp.append(comp_headers)
    for idx, row in compat_df.iterrows():
        ws_comp.append([idx] + [int(x) for x in row.tolist()])

    for ws_current in [ws, ws_mat, ws_mach, ws_parts, ws_comp]:
        for col in ws_current.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws_current.column_dimensions[col_letter].width = min(max_len + 2, 25)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def style_figure(fig, title, x_title, y_title, height):
    fig.update_layout(
        title=dict(text=title, font=dict(color="#111111", size=20)),
        xaxis_title=dict(text=x_title, font=dict(color="#111111", size=15)),
        yaxis_title=dict(text=y_title, font=dict(color="#111111", size=15)),
        height=height,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(color="#111111", size=13),
        legend=dict(font=dict(color="#111111", size=12), bgcolor="rgba(255,255,255,0.9)"),
        legend_title_text="",
        margin=dict(l=50, r=30, t=80, b=50),
    )
    fig.update_xaxes(showgrid=True, gridcolor="#CFCFCF", zeroline=True, zerolinecolor="#B0B0B0", tickfont=dict(color="#111111", size=12))
    fig.update_yaxes(showgrid=True, gridcolor="#ECECEC", tickfont=dict(color="#111111", size=12))
    return fig


def make_gantt_chart(gantt_df):
    if gantt_df.empty:
        return None
    machine_order = list(reversed(sorted(gantt_df["Machine"].unique().tolist())))
    fig = go.Figure()
    colors = {}
    palette = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#17becf"]
    materials = gantt_df["Material"].unique().tolist()
    for i, mat in enumerate(materials):
        colors[mat] = palette[i % len(palette)]

    for _, row in gantt_df.iterrows():
        fig.add_trace(
            go.Bar(
                x=[row["Duration"]],
                y=[row["Machine"]],
                base=[row["Start"]],
                orientation="h",
                marker_color=colors[row["Material"]],
                marker_line_color="#444444",
                marker_line_width=1,
                name=row["Material"],
                text=row["Batch"],
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="white", size=12),
                hovertemplate=f"Machine: {row['Machine']}<br>Batch: {row['Batch']}<br>Matériau: {row['Material']}<br>Début: {row['Start']:.2f}<br>Fin: {row['Start'] + row['Duration']:.2f}<extra></extra>",
                showlegend=False,
            )
        )

    for mat in materials:
        fig.add_trace(go.Bar(x=[0], y=[machine_order[0]], marker_color=colors[mat], name=mat, visible="legendonly", orientation="h"))

    fig.update_yaxes(categoryorder="array", categoryarray=machine_order)
    fig.update_layout(barmode="overlay")
    return style_figure(fig, "Gantt de production par machine", "Temps de production", "Machine", 460)


def make_batch_comp_chart(comp_df, batch_df):
    if comp_df.empty:
        return None
    batch_order = list(reversed(batch_df.sort_values(["machine", "batch_index"])["batch_id"].tolist()))
    fig = go.Figure()
    colors = {}
    palette = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#17becf"]
    materials = comp_df["Material"].unique().tolist()
    for i, mat in enumerate(materials):
        colors[mat] = palette[i % len(palette)]

    for _, row in comp_df.iterrows():
        fig.add_trace(
            go.Bar(
                x=[row["Duration"]],
                y=[row["Batch"]],
                base=[row["Start"]],
                orientation="h",
                marker_color=colors[row["Material"]],
                marker_line_color="#444444",
                marker_line_width=1,
                text=row["Part"],
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="white", size=11),
                hovertemplate=f"Batch: {row['Batch']}<br>Pièce: {row['Part']}<br>Matériau: {row['Material']}<br>Début segment: {row['Start']:.2f}<br>Fin segment: {row['Start'] + row['Duration']:.2f}<extra></extra>",
                showlegend=False,
            )
        )
    for mat in materials:
        fig.add_trace(go.Bar(x=[0], y=[batch_order[0]], marker_color=colors[mat], name=mat, visible="legendonly", orientation="h"))

    fig.update_yaxes(categoryorder="array", categoryarray=batch_order)
    fig.update_layout(barmode="overlay")
    return style_figure(fig, "Composition des batches", "Aire cumulée dans le batch", "Batch", max(360, 70 * len(batch_order)))


def make_load_chart(machine_summary_df):
    if machine_summary_df.empty:
        return None
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=machine_summary_df["machine"],
            y=machine_summary_df["charge_totale"],
            text=machine_summary_df["charge_totale"],
            textposition="outside",
            marker_color="#4C78A8",
            marker_line_color="#333333",
            marker_line_width=1,
            hovertemplate="Machine: %{x}<br>Charge totale: %{y:.2f}<extra></extra>",
        )
    )
    return style_figure(fig, "Charge totale par machine", "Machine", "Temps total affecté", 380)


def make_material_chart(batch_df):
    if batch_df.empty:
        return None
    tmp = batch_df.groupby("material", as_index=False)["nb_pieces"].sum()
    fig = go.Figure(
        data=[
            go.Pie(
                labels=tmp["material"],
                values=tmp["nb_pieces"],
                hole=0.35,
                textinfo="label+percent",
                textfont=dict(color="#111111", size=13),
            )
        ]
    )
    fig.update_layout(
        title=dict(text="Répartition des pièces par matériau", font=dict(color="#111111", size=20)),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111111", size=13),
        legend=dict(font=dict(color="#111111", size=12), bgcolor="rgba(255,255,255,0.9)"),
        height=380,
        margin=dict(l=30, r=30, t=70, b=30),
    )
    return fig


init_state()
sync_all()

st.title("Application interactive de scheduling en fabrication additive")
st.markdown(
    "<div style='text-align: right; font-size:14px; color:gray;'>"
    "Fait par Émile Fortin et Nicholas Ménard"
    "</div>",
    unsafe_allow_html=True,
)
st.caption("Application préparée pour la présentation du projet au professeur.")

with st.expander("Définition des données et variables de l'instance", expanded=False):
    st.markdown("""
**Pièces**
- `part_id` : identifiant de la pièce
- `h` : hauteur de la pièce
- `a` : aire occupée par la pièce
- `v` : volume de la pièce
- `material_name` : nom du matériau de la pièce

**Machines**
- `machine_id` : identifiant de la machine
- `VT` : coefficient lié au volume dans le temps de traitement
- `HT` : coefficient lié à la hauteur maximale du batch
- `SET` : temps de setup
- `MA` : capacité maximale en aire
- `MH` : hauteur maximale imprimable

**Compatibilité**
- `1` : la machine peut traiter ce matériau
- `0` : la machine ne peut pas traiter ce matériau

**Planification**
- `jn` : nombre maximal de batches par machine
- `CMAX` : durée totale du planning
    """)

st.subheader("1) Paramètres globaux")
st.session_state.jn = st.number_input("Nombre max de batches par machine (jn)", 1, 30, int(st.session_state.jn))

st.subheader("2) Génération des données")
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("**Matériaux**")
    nb_materials = st.number_input("Nombre de matériaux", 1, 10, len(st.session_state.materials_df))
    if st.button("Générer les matériaux", use_container_width=True):
        generate_materials(int(nb_materials))

    st.markdown("**Compatibilité**")
    st.caption("Remplit automatiquement la matrice machine-matériau avec des 0/1 cohérents.")
    if st.button("Générer la compatibilité", use_container_width=True):
        generate_compatibility()

with col2:
    st.markdown("**Machines**")
    nb_machines = st.number_input("Nombre de machines", 1, 10, len(st.session_state.machines_df))
    ma_min = st.number_input("MA min", 10.0, 10000.0, 700.0)
    ma_max = st.number_input("MA max", 10.0, 10000.0, 1400.0)
    mh_min = st.number_input("MH min", 1.0, 500.0, 28.0)
    mh_max = st.number_input("MH max", 1.0, 500.0, 45.0)
    if st.button("Générer les machines", use_container_width=True):
        generate_machines(int(nb_machines), float(ma_min), float(ma_max), float(mh_min), float(mh_max))

with col3:
    st.markdown("**Pièces**")
    nb_parts = st.number_input("Nombre de pièces", 1, 300, len(st.session_state.parts_df))
    h_min = st.number_input("h min", 0.1, 1000.0, 2.0)
    h_max = st.number_input("h max", 0.1, 1000.0, 30.0)
    a_min = st.number_input("a min", 1.0, 100000.0, 20.0)
    a_max = st.number_input("a max", 1.0, 100000.0, 500.0)
    v_min = st.number_input("v min", 1.0, 100000.0, 80.0)
    v_max = st.number_input("v max", 1.0, 100000.0, 3200.0)
    if st.button("Générer les pièces", use_container_width=True):
        generate_parts(int(nb_parts), float(h_min), float(h_max), float(a_min), float(a_max), float(v_min), float(v_max))

st.subheader("3) Données modifiables")
tab1, tab2, tab3, tab4 = st.tabs(["Matériaux", "Machines", "Pièces", "Compatibilité"])

with tab1:
    materials_edited = st.data_editor(
        st.session_state.materials_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "material_id": st.column_config.NumberColumn("ID", disabled=True),
            "material_name": st.column_config.TextColumn("Nom du matériau"),
        },
        key="materials_editor",
    )
    st.session_state.materials_df = materials_edited.copy()
    sync_all()

with tab2:
    machines_edited = st.data_editor(
        st.session_state.machines_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "machine_id": st.column_config.NumberColumn("ID machine", disabled=True),
            "VT": st.column_config.NumberColumn("VT", format="%.6f"),
            "HT": st.column_config.NumberColumn("HT", format="%.3f"),
            "SET": st.column_config.NumberColumn("SET", format="%.3f"),
            "MA": st.column_config.NumberColumn("MA", format="%.2f"),
            "MH": st.column_config.NumberColumn("MH", format="%.2f"),
        },
        key="machines_editor",
    )
    st.session_state.machines_df = machines_edited.copy()
    sync_all()

with tab3:
    part_material_options = st.session_state.materials_df["material_name"].tolist() or ["PLA"]
    parts_edited = st.data_editor(
        st.session_state.parts_df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "part_id": st.column_config.NumberColumn("ID pièce", disabled=True),
            "h": st.column_config.NumberColumn("h", format="%.2f"),
            "a": st.column_config.NumberColumn("a", format="%.2f"),
            "v": st.column_config.NumberColumn("v", format="%.2f"),
            "material_name": st.column_config.SelectboxColumn("Matériau", options=part_material_options),
        },
        key="parts_editor",
    )
    st.session_state.parts_df = parts_edited.copy()
    sync_all()

with tab4:
    compat_editor = st.data_editor(
        st.session_state.compat_df,
        use_container_width=True,
        hide_index=False,
        key="compat_editor",
    )
    st.session_state.compat_df = compat_editor.copy()
    sync_all()

st.subheader("4) Diagnostic avant résolution")
legend_cols = st.columns(3)
legend_cols[0].markdown("🟥 **Erreur bloquante** : impossible de construire une solution cohérente")
legend_cols[1].markdown("🟨 **Avertissement** : solution possible mais attention")
legend_cols[2].markdown("🟩 **Validation** : aucun problème bloquant détecté")

errors, warnings = validate_dataframes(
    st.session_state.parts_df,
    st.session_state.machines_df,
    st.session_state.compat_df,
    int(st.session_state.jn),
)

if errors:
    for err in errors:
        st.error(err)
else:
    st.success("Aucune erreur bloquante détectée dans les données.")

if warnings:
    for warn in warnings[:15]:
        st.warning(warn)

dat_text = build_opl_dat(
    st.session_state.parts_df,
    st.session_state.machines_df,
    st.session_state.compat_df,
    st.session_state.materials_df,
    int(st.session_state.jn),
)

excel_file = build_excel_export(
    st.session_state.parts_df,
    st.session_state.machines_df,
    st.session_state.compat_df,
    st.session_state.materials_df,
    int(st.session_state.jn),
)

c_download1, c_download2, c_run = st.columns([1, 1, 1])
with c_download1:
    st.download_button(
        "Télécharger le fichier .dat OPL",
        data=dat_text,
        file_name="ParallelNonIdentical_generated.dat",
        mime="text/plain",
        use_container_width=True,
    )
with c_download2:
    st.download_button(
        "Télécharger le fichier Excel",
        data=excel_file,
        file_name="am_scheduling_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with c_run:
    run_planning = st.button("5) Construire le planning", use_container_width=True)

if run_planning:
    batch_df, gantt_df, comp_df, machine_summary_df, cmax, unassigned = schedule_heuristic(
        st.session_state.parts_df,
        st.session_state.machines_df,
        st.session_state.compat_df,
        int(st.session_state.jn),
    )

    st.subheader("5) Résultats du planning")
    k1, k2, k3 = st.columns(3)
    k1.metric("CMAX", cmax)
    k2.metric("Nombre total de batches utilisés", 0 if batch_df.empty else len(batch_df))
    k3.metric("Pièces non affectées", len(unassigned))

    if unassigned:
        st.error("Pièces non affectées : " + ", ".join(f"P{i}" for i in unassigned))

    st.markdown("**Résumé par machine**")
    if not machine_summary_df.empty:
        st.dataframe(machine_summary_df, use_container_width=True, hide_index=True)

    st.markdown("**Détail complet des batches**")
    if not batch_df.empty:
        st.dataframe(batch_df, use_container_width=True, hide_index=True)

    gantt_fig = make_gantt_chart(gantt_df)
    if gantt_fig is not None:
        st.plotly_chart(gantt_fig, use_container_width=True)

    comp_fig = make_batch_comp_chart(comp_df, batch_df)
    if comp_fig is not None:
        st.plotly_chart(comp_fig, use_container_width=True)

    extra1, extra2 = st.columns(2)
    with extra1:
        load_fig = make_load_chart(machine_summary_df)
        if load_fig is not None:
            st.plotly_chart(load_fig, use_container_width=True)
    with extra2:
        mat_fig = make_material_chart(batch_df)
        if mat_fig is not None:
            st.plotly_chart(mat_fig, use_container_width=True)

with st.expander("Que fait cette application ?", expanded=False):
    st.write(
        """
Cette application permet de :
- générer des matériaux, des machines, des pièces et une matrice de compatibilité ;
- modifier directement toutes les données dans les grilles ;
- détecter les incohérences avant la résolution ;
- construire un planning heuristique ;
- visualiser le planning global et la composition détaillée de chaque batch ;
- ajouter des graphiques complémentaires sur la charge des machines et la répartition des matériaux ;
- exporter les données en fichier `.dat` pour OPL et en fichier Excel.
        """
    )
